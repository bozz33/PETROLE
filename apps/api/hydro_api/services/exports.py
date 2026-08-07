"""Exports tabulaires des résultats de calcul.

Le MVP exige que l'ingénieur puisse récupérer ses résultats en PDF, XLSX, CSV et
JSON. Le PDF est produit par le module de rapports ; ce module couvre les trois
formats tabulaires, à partir du résultat persisté et sans recalcul.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook

from hydro_shared.hashing import canonical_json

#: Sections exportables et colonnes retenues, dans l'ordre d'affichage.
EXPORT_SECTIONS: dict[str, tuple[str, ...]] = {
    "profile": (
        "chainage_m",
        "elevation_m",
        "pressure_pa",
        "hydraulic_grade_m",
        "flow_m3_s",
        "velocity_m_s",
        "below_vapor_pressure",
        "gravity_zone",
    ),
    "segments": (
        "segment_id",
        "label",
        "flow_m3_s",
        "velocity_m_s",
        "reynolds",
        "friction_factor",
        "friction_model",
        "friction_head_loss_m",
        "minor_head_loss_m",
        "total_head_loss_m",
        "elevation_change_m",
        "inlet_pressure_pa",
        "outlet_pressure_pa",
        "min_pressure_pa",
        "max_pressure_pa",
        "maop_margin_pa",
    ),
    "stations": (
        "station_id",
        "name",
        "chainage_m",
        "elevation_m",
        "in_service",
        "bypassed",
        "flow_m3_s",
        "suction_pressure_pa",
        "discharge_pressure_pa",
        "differential_pressure_pa",
        "head_m",
        "hydraulic_power_w",
        "absorbed_power_w",
        "efficiency",
        "active_pump_count",
    ),
    "pumps": (
        "pump_id",
        "label",
        "station_id",
        "running",
        "flow_m3_s",
        "head_m",
        "speed_ratio",
        "efficiency",
        "hydraulic_power_w",
        "absorbed_power_w",
        "npsh_required_m",
        "npsh_available_m",
        "npsh_margin_m",
        "within_curve_domain",
    ),
}


def section_rows(result_payload: dict[str, Any], section: str) -> list[dict[str, Any]]:
    """Extrait les lignes d'une section du résultat, colonnes normalisées.

    Les pompes ne sont pas une section du résultat : elles sont imbriquées dans
    les stations. Elles sont remontées ici pour rester exportables telles quelles.
    """

    if section not in EXPORT_SECTIONS:
        raise ValueError(f"Section exportable inconnue : {section}.")

    if section == "pumps":
        raw: list[Any] = []
        for station in result_payload.get("stations") or []:
            if isinstance(station, dict):
                raw.extend(station.get("pumps") or [])
    else:
        raw = list(result_payload.get(section) or [])

    columns = EXPORT_SECTIONS[section]
    return [
        {column: item.get(column) for column in columns} for item in raw if isinstance(item, dict)
    ]


def to_csv(rows: list[dict[str, Any]], section: str) -> bytes:
    """Sérialise une section en CSV, séparateur point-virgule et en-tête stable."""

    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=list(EXPORT_SECTIONS[section]),
        delimiter=";",
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _csv_value(value) for key, value in row.items()})
    # Le BOM permet à un tableur de reconnaître l'encodage sans réglage manuel.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_xlsx(sections: dict[str, list[dict[str, Any]]]) -> bytes:
    """Assemble un classeur, une feuille par section exportée."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    for section, rows in sections.items():
        sheet = workbook.create_sheet(title=section[:31])
        columns = list(EXPORT_SECTIONS[section])
        sheet.append(columns)
        for row in rows:
            sheet.append([_cell_value(row.get(column)) for column in columns])
    if not workbook.sheetnames:
        workbook.create_sheet(title="vide")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_json(payload: dict[str, Any]) -> bytes:
    """Sérialise un export en JSON canonique, stable d'une exécution à l'autre."""

    return canonical_json(payload).encode("utf-8")


def _csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "oui" if value else "non"
    return str(value)


def _cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "oui" if value else "non"
    if value is None or isinstance(value, int | float | str):
        return value
    return str(value)


__all__ = ["EXPORT_SECTIONS", "section_rows", "to_csv", "to_json", "to_xlsx"]
